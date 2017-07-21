import json
from collections import OrderedDict

import os

from openpyxl import Workbook


def getDoc(path):
    with open(path,'r',encoding='utf-8') as f:
        return OrderedDict(json.load(f))
def saveAsExcel(dict,save_path):
    wb=Workbook()
    ws=wb.create_sheet('student',0)
    row=1
    for k,v in student_dict.items():
        ws.cell(row=row, column=1, value=k)
        col=2
        for item in v:
            ws.cell(row=row, column=col, value=item)
            col=col+1
        row=row+1
        print(k,v)
    wb.save(save_path)
if __name__=='__main__':
    input_path=os.getcwd()+os.path.sep+'student.txt'
    output_path=os.getcwd()+os.path.sep+'students.xlsx'
    student_dict=getDoc(input_path)
    saveAsExcel(student_dict,output_path)

